"""
广州美食数据库采集模板生成工具 V1.0

功能:
1. 自动生成Excel采集模板
2. 创建餐厅信息表
3. 创建菜品信息表
4. 创建推荐关系表
5. 创建早餐/午餐/晚餐数据规划表

Python:
3.12+

依赖:
openpyxl
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment
)
from openpyxl.utils import get_column_letter

# =====================================================
# 路径
# =====================================================


BASE_DIR = Path(__file__).resolve().parent.parent

OUTPUT_DIR = (

        BASE_DIR

        /

        "data"

        /

        "source"

)

OUTPUT_DIR.mkdir(
    parents=True,
    exist_ok=True
)

OUTPUT_FILE = (

        OUTPUT_DIR

        /

        "food_collection_template.xlsx"

)


# =====================================================
# Excel格式
# =====================================================


def format_sheet(ws):
    """
    设置Excel格式
    """

    for cell in ws[1]:
        cell.font = Font(
            bold=True
        )

        cell.fill = PatternFill(
            "solid",
            fgColor="D9EAF7"
        )

        cell.alignment = Alignment(
            horizontal="center"
        )

    for column in ws.columns:

        max_length = 0

        for cell in column:

            if cell.value:
                max_length = max(

                    max_length,

                    len(
                        str(cell.value)
                    )

                )

        ws.column_dimensions[

            get_column_letter(
                column[0].column
            )

        ].width = min(

            max_length + 5,

            30

        )


# =====================================================
# 创建餐厅模板
# =====================================================


def create_restaurant_sheet(wb):
    ws = wb.create_sheet(

        "restaurants"

    )

    headers = [

        "id",

        "name",

        "district",

        "area",

        "category",

        "price",

        "rating",

        "couple_score",

        "environment_score",

        "taste_score",

        "business_hours",

        "address",

        "latitude",

        "longitude",

        "map_url",

        "remark"

    ]

    ws.append(headers)

    ws.append([

        10001,

        "示例：陶陶居",

        "越秀",

        "北京路",

        "粤菜",

        100,

        4.8,

        5,

        5,

        5,

        "08:00-22:00",

        "广州越秀区",

        "",

        "",

        "",

        ""

    ])

    format_sheet(ws)


# =====================================================
# 创建菜品模板
# =====================================================


def create_dish_sheet(wb):
    ws = wb.create_sheet(

        "dishes"

    )

    headers = [

        "id",

        "name",

        "meal",

        "category",

        "type",

        "tags",

        "description",

        "remark"

    ]

    ws.append(headers)

    ws.append([

        20001,

        "虾饺皇",

        "早茶|早餐",

        "粤菜",

        "点心",

        "经典,情侣,老字号",

        "广州传统早茶代表",

        ""

    ])

    format_sheet(ws)


# =====================================================
# 创建关系模板
# =====================================================


def create_relation_sheet(wb):
    ws = wb.create_sheet(

        "food_relation"

    )

    headers = [

        "id",

        "restaurant_id",

        "dish_id",

        "meal",

        "scene",

        "tags",

        "reason"

    ]

    ws.append(headers)

    ws.append([

        30001,

        10001,

        20001,

        "早餐",

        "情侣约会",

        "经典,推荐",

        "适合第一次体验广州早茶"

    ])

    format_sheet(ws)


# =====================================================
# 创建数据规划表
# =====================================================


def create_plan_sheet(wb):
    ws = wb.create_sheet(

        "data_plan"

    )

    headers = [

        "meal",

        "target_count",

        "category",

        "status",

        "remark"

    ]

    ws.append(headers)

    plans = [

        [

            "早餐",

            100,

            "早茶/肠粉/粥粉面",

            "待收集",

            ""

        ],

        [

            "午餐",

            150,

            "粤菜/烧腊/快餐",

            "待收集",

            ""

        ],

        [

            "晚餐",

            150,

            "情侣餐厅/特色餐厅",

            "待收集",

            ""

        ]

    ]

    for item in plans:
        ws.append(item)

    format_sheet(ws)


# =====================================================
# 主程序
# =====================================================


def main():
    print("=" * 50)

    print(
        "广州美食数据采集模板生成工具 V1.0"
    )

    print("=" * 50)

    wb = Workbook()

    # 删除默认Sheet

    default = wb.active

    wb.remove(default)

    create_restaurant_sheet(wb)

    create_dish_sheet(wb)

    create_relation_sheet(wb)

    create_plan_sheet(wb)

    wb.save(

        OUTPUT_FILE

    )

    print(

        f"✅ 模板生成完成:"
        f"\n{OUTPUT_FILE}"

    )


if __name__ == "__main__":
    main()