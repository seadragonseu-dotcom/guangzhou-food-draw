"""
广州美食数据库

300家真实餐厅采集模板生成工具 V1.0

输出:
source/
└── restaurant_source.xlsx


Python:
3.12+
"""


from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter



# =====================================================
# Path
# =====================================================


BASE_DIR = Path(__file__).resolve().parent.parent


SOURCE_DIR = BASE_DIR / "source"


OUTPUT_FILE = (
    SOURCE_DIR /
    "restaurant_source.xlsx"
)



# =====================================================
# Config
# =====================================================


TOTAL_COUNT = 300



HEADERS = [

    "id",

    "name",

    "district",

    "area",

    "category",

    "price",

    "rating",

    "couple_score",

    "taste_score",

    "environment_score",

    "business_hours",

    "address",

    "latitude",

    "longitude",

    "map_url",

    "source",

    "remark",

    "real_flag",

    "verify_date"

]



# =====================================================
# Excel Style
# =====================================================


HEADER_FILL = PatternFill(

    "solid",

    fgColor="FFD966"

)



HEADER_FONT = Font(

    bold=True

)



CENTER = Alignment(

    horizontal="center",

    vertical="center"

)



# =====================================================
# Create
# =====================================================


def create_template():



    print("=" * 60)

    print(
        "广州300家真实餐厅采集模板生成工具 V1.0"
    )

    print("=" * 60)



    SOURCE_DIR.mkdir(

        parents=True,

        exist_ok=True

    )



    wb = Workbook()



    ws = wb.active



    ws.title = "restaurants"



    # -----------------------------
    # Header
    # -----------------------------


    ws.append(

        HEADERS

    )



    for cell in ws[1]:


        cell.fill = HEADER_FILL

        cell.font = HEADER_FONT

        cell.alignment = CENTER



    # -----------------------------
    # Empty rows
    # -----------------------------


    for i in range(

        1,

        TOTAL_COUNT + 1

    ):


        ws.append(

            [

                i,

                "",

                "",

                "",

                "",

                "",

                "",

                "",

                "",

                "",

                "",

                "",

                "",

                "",

                "",

                "",

                "",

                1,

                ""

            ]

        )



    # -----------------------------
    # Freeze
    # -----------------------------


    ws.freeze_panes = "A2"



    # Filter

    ws.auto_filter.ref = (

        f"A1:{get_column_letter(len(HEADERS))}"

        f"{TOTAL_COUNT + 1}"

    )



    # -----------------------------
    # Column Width
    # -----------------------------


    widths = {


        "A":8,

        "B":20,

        "C":12,

        "D":15,

        "E":15,

        "F":10,

        "G":10,

        "H":12,

        "I":12,

        "J":15,

        "K":18,

        "L":35,

        "M":14,

        "N":14,

        "O":45,

        "P":25,

        "Q":20,

        "R":10,

        "S":15

    }



    for col,width in widths.items():


        ws.column_dimensions[col].width = width



    # -----------------------------
    # Data Validation
    # -----------------------------


    district_validation = DataValidation(

        type="list",

        formula1=(

            '"越秀,荔湾,海珠,天河,白云,番禺,黄埔,花都,南沙,增城,从化"'

        )

    )


    category_validation = DataValidation(

        type="list",

        formula1=(

            '"粤菜,早茶,烧腊,潮汕菜,火锅,小吃,甜品,西餐,日料,其他"'

        )

    )



    source_validation = DataValidation(

        type="list",

        formula1=(

            '"美团,大众点评,高德地图,官网,人工采集"'

        )

    )



    flag_validation = DataValidation(

        type="list",

        formula1='"1,0"'

    )



    ws.add_data_validation(

        district_validation

    )


    ws.add_data_validation(

        category_validation

    )


    ws.add_data_validation(

        source_validation

    )


    ws.add_data_validation(

        flag_validation

    )



    district_validation.add(

        f"C2:C{TOTAL_COUNT+1}"

    )


    category_validation.add(

        f"E2:E{TOTAL_COUNT+1}"

    )


    source_validation.add(

        f"P2:P{TOTAL_COUNT+1}"

    )


    flag_validation.add(

        f"R2:R{TOTAL_COUNT+1}"

    )



    # -----------------------------
    # Save
    # -----------------------------


    wb.save(

        OUTPUT_FILE

    )



    print()

    print(

        f"生成餐厅数量:{TOTAL_COUNT}"

    )



    print()

    print(

        "输出文件:"

    )


    print(

        OUTPUT_FILE

    )



    print()

    print(

        "更新时间:",

        datetime.now()

    )



    print()

    print(

        "模板生成完成"

    )





if __name__ == "__main__":


    create_template()